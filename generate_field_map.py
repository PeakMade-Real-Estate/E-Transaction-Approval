"""
generate_field_map.py
Generates 'E-Transaction Field Map.xlsx' — run once: python generate_field_map.py

5 sheets:
  1. Master Field Reference   — all fields, all pages
  2. Intake Form              — fields written when user submits
  3. Dashboard                — fields read on /dashboard
  4. Request Detail           — fields read/written on /dashboard/request/<id>
  5. Confirmation             — fields read on /confirmation/<id>
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY   = "1E3A5F"
WHITE  = "FFFFFF"
SEC_A  = "D6E4F7"   # light blue
SEC_B  = "FFF3CD"   # light yellow
SEC_C  = "FDE8D0"   # light orange
SEC_D  = "EDE0F5"   # light purple
SEC_E  = "D6F0E0"   # light green
SEC_SYS= "E8E8E8"   # light gray
ROW_ALT= "F4F4F4"

SECTION_FILL = {
    "A": SEC_A, "B": SEC_B, "C": SEC_C,
    "D": SEC_D, "E": SEC_E,
    "System": SEC_SYS, "Sidebar": SEC_SYS,
}

RW_STYLE = {
    "Write":      ("D6F0E0", "1A6B35", True),
    "Read/Write": ("EDE0F5", "4B0082", True),
    "Computed":   ("FFF3CD", "7B5A00", True),
    "Read":       (None,     NAVY,    False),
    "Mock":       (SEC_SYS,  "555555", False),
    "N/A":        (SEC_SYS,  "555555", False),
}


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _font(bold=False, color="000000", size=9):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def write_sheet(wb, title, headers, rows, col_widths,
                section_col=None, rw_col=None):
    ws = wb.create_sheet(title)

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = _font(bold=True, color=WHITE, size=10)
        cell.fill      = _fill(NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for row_data in rows:
        ws.append(list(row_data))
        r = ws.max_row

        # Row background — section color or alternating gray
        if section_col is not None and len(row_data) > section_col:
            sec = str(row_data[section_col])
            row_color = SECTION_FILL.get(sec, ROW_ALT if r % 2 == 0 else WHITE)
        else:
            row_color = ROW_ALT if r % 2 == 0 else WHITE

        for cell in ws[r]:
            cell.fill      = _fill(row_color)
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
            cell.font      = _font()

        # Highlight Read/Write cell
        if rw_col is not None and len(row_data) > rw_col:
            val  = str(row_data[rw_col]).strip()
            cell = ws.cell(row=r, column=rw_col + 1)
            if val in RW_STYLE:
                bg, fg, bld = RW_STYLE[val]
                if bg:
                    cell.fill = _fill(bg)
                cell.font = _font(bold=bld, color=fg)

    # Column widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Master Field Reference
# ══════════════════════════════════════════════════════════════════════════════

MASTER_HEADERS = [
    "Field Key", "Display Label", "Python Type", "Section",
    "Intake", "Dashboard", "Detail", "Confirmation",
    "Source / Origin", "SQL Operation", "Notes / Future Integration Tag",
]

#  Col:  0-key  1-label  2-type  3-section  4-intake  5-dash  6-detail  7-conf  8-source  9-sql_op  10-notes
MASTER_ROWS = [
    ("request_id",               "Request ID",                         "str",   "System", "Computed", "Y","Y","Y", "Generated",  "Write",      "TXN-YYYY-###  (random 3-digit suffix)"),
    ("submitted_date",           "Submitted Date",                     "str",   "System", "Computed", "Y","Y","Y", "Generated",  "Write",      "Set to today on intake submit (YYYY-MM-DD)"),
    ("request_type",             "Request Type",                       "str",   "A",      "Y",        "Y","Y","Y", "Form",       "Write",      "ACH / Wire / EFT / Intra Bank Transfer"),
    ("property_dept",            "Property / Department",              "str",   "A",      "Y",        "Y","Y","Y", "Form",       "Write",      "Free text"),
    ("property_code",            "Entity / Property Code",             "str",   "A",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("prepared_by",              "Prepared By",                        "str",   "A",      "Y",        "Y","Y","Y", "Form",       "Write",      "[AUTH] Replace with Azure AD display name"),
    ("prepared_date",            "Prepared Date",                      "str",   "A",      "Y",        "N","Y","N", "Form",       "Write",      "Auto-set by JS to today"),
    ("approver",                 "Approver",                           "str",   "A",      "Y",        "N","Y","N", "Form",       "Write",      "[AUTH] Will become AD user picker"),
    ("controller",               "Controller",                         "str",   "A",      "Y",        "N","Y","N", "Form",       "Write",      "[AUTH] Will become AD user picker"),
    ("treasury_service_date",    "Treasury Service Date",              "str",   "A",      "Y",        "N","Y","Y", "Form",       "Write",      ""),
    ("instructions_previously_used", "Previously Used Instructions",  "bool",  "A",      "Y",        "N","Y","N", "Form",       "Write",      "Checkbox; controls last_used_date visibility"),
    ("last_used_date",           "Last Date Used",                     "str",   "A",      "Y",        "N","Y","N", "Form",       "Write",      "Required only when instructions_previously_used=True"),
    ("urgent",                   "Urgent?",                            "bool",  "A",      "Y",        "Y","Y","Y", "Form",       "Write",      "Radio yes/no; drives urgency banner and flag badge"),
    ("urgency_reason",           "Urgency Reason",                     "str",   "A",      "Y",        "N","Y","Y", "Form",       "Write",      "Conditional — shown when urgent=True"),
    ("verbal_confirmed",         "Instructions Verbally Confirmed",    "bool",  "B",      "Y",        "N","Y","N", "Form",       "Write",      "Checkbox"),
    ("verbal_confirmed_with",    "Verbally Confirmed With",            "str",   "B",      "Computed", "N","Y","N", "Composite",  "Write",      "Joined from two checkboxes: Known Contact / Requesting Person"),
    ("verbal_contact_name",      "Verbal Contact Name",                "str",   "B",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("verbal_confirm_datetime",  "Verbal Confirmation Date/Time",      "str",   "B",      "Y",        "N","Y","N", "Form",       "Write",      "datetime-local input"),
    ("avs_score",                "AVS Score",                          "str",   "B",      "Y",        "N","Y","N", "Form",       "Write",      "Stored as str; UI warns if cast to int <90. [BANKING]"),
    ("external_source",          "Instructions from External Source",  "bool",  "B",      "Y",        "N","Y","N", "Form",       "Write",      "Checkbox"),
    ("internal_doc_not_used",    "Internal Document NOT Used",         "bool",  "B",      "Y",        "N","Y","N", "Form",       "Write",      "Checkbox"),
    ("attachments.validation_evidence",  "Validation Evidence File",   "str",   "B",      "Y",        "N","Y","N", "File upload","Write",      "Filename only stored. [UPLOAD]"),
    ("orig_bank_name",           "Originating Bank Name",              "str",   "C",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("orig_account_name",        "Originating Account Name",           "str",   "C",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("orig_account_number",      "Originating Account Number",         "str",   "C",      "Y",        "N","Y","N", "Form",       "Write",      "[RBAC] Masked on detail page; reveal toggle"),
    ("orig_routing_number",      "Originating Routing Number",         "str",   "C",      "Y",        "N","Y","N", "Form",       "Write",      "[RBAC] Masked on detail page; reveal toggle"),
    ("orig_bank_contact",        "Originating Bank Contact",           "str",   "C",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("notes_orig",               "Notes — Originating Bank",           "str",   "C",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("recv_payee_name",          "Recipient / Payee Name",             "str",   "D",      "Y",        "N","Y","Y", "Form",       "Write",      ""),
    ("recv_bank_name",           "Receiving Bank Name",                "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("recv_account_name",        "Receiving Account Name",             "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("recv_account_number",      "Receiving Account Number",           "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      "[RBAC] Masked; reveal toggle"),
    ("recv_routing_number",      "Receiving Routing Number",           "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      "[RBAC] Masked; reveal toggle"),
    ("recv_bank_address",        "Bank / Beneficiary Address",         "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      "Required when request_type = Wire"),
    ("recv_contact_name",        "Receiving Contact Name",             "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("recv_contact_email",       "Receiving Contact Email",            "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("recv_contact_phone",       "Receiving Contact Phone",            "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("notes_recv",               "Notes — Receiving Bank",             "str",   "D",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("attachments.wire_ach_instructions", "External ACH/Wire Instructions File", "str", "D", "Y", "N","Y","N", "File upload", "Write",      "Filename only. [UPLOAD]; Missing badge if absent"),
    ("amount",                   "Amount",                             "float", "E",      "Y",        "Y","Y","Y", "Form",       "Write",      "Comma-stripped before float(); formatted via |currency filter"),
    ("currency",                 "Currency",                           "str",   "E",      "Y",        "N","Y","Y", "Form",       "Write",      "Default: USD"),
    ("payment_purpose",          "Payment Purpose / Description",      "str",   "E",      "Y",        "N","Y","N", "Form",       "Write",      ""),
    ("attachments.payment_support", "Payment Support File",            "str",   "E",      "Y",        "N","Y","N", "File upload","Write",      "Filename only. [UPLOAD]; Missing badge if absent"),
    ("approval_tier",            "Required Approval Tier",             "str",   "System", "Computed", "N","Y","Y", "Computed",   "Write",      "get_approval_tier(amount): SAM / Controller / VP / CFO"),
    ("status",                   "Status",                             "str",   "System", "Computed", "Y","Y","N", "Generated",  "Read/Write", "Initial: 'Submitted'; updated by request_action(). [WORKFLOW]"),
    ("over_1m",                  "Over $1M Flag",                      "bool",  "System", "Computed", "Y","Y","Y", "Computed",   "Write",      "amount > 1,000,000"),
    ("assigned_approver",        "Assigned Approver",                  "str",   "System", "Computed", "Y","N","N", "Generated",  "Read/Write", "Initial: 'Pending Assignment'. [WORKFLOW]"),
    ("days_pending",             "Days Pending",                       "int",   "System", "Computed", "Y","N","N", "Generated",  "Read/Write", "Initial: 0. [WORKFLOW] Should compute from submitted_date"),
    ("timeline",                 "Approval Timeline",                  "list",  "System", "Computed", "N","Y","N", "Generated",  "Read/Write", "List of {date, event, actor, status, type}; appended on each action"),
    ("comments",                 "Comments / Notes",                   "list",  "System", "Computed", "N","Y","N", "Generated",  "Read/Write", "List of {author, date, text}; appended via request_action()"),
    ("extra_attachments",        "Additional Attachments",             "list",  "System", "N",        "N","Y","N", "POST /attach","Read/Write", "List of {filename, description, uploaded_by, date}"),
    ("docs_checklist.treasury_template",  "Treasury Template",         "bool",  "System", "Mock",     "N","Y","N", "Mock only",  "Read",       ""),
    ("docs_checklist.amount_backup",      "Amount Backup Documentation","bool", "System", "Mock",     "N","Y","N", "Mock only",  "Read",       ""),
    ("docs_checklist.auth_to_move",       "Authorization to Move Funds","bool", "System", "Mock",     "N","Y","N", "Mock only",  "Read",       ""),
    ("docs_checklist.exec_approval",      "Executive / Officer Approval","bool","System", "Mock",     "N","Y","N", "Mock only",  "Read",       "Required when over_1m=True"),
    ("docs_checklist.external_wire_ach",  "External Wire/ACH Instructions","bool","System","Mock",    "N","Y","N", "Mock only",  "Read",       ""),
    ("docs_checklist.wf_avs_screenshot",  "WF AVS Screenshot",         "bool",  "B",      "Mock",     "N","Y","N", "Mock only",  "Read",       "Shown in Section B: AVS Validation Evidence column"),
    ("docs_checklist.final_pdf",          "Final PDF",                  "bool", "System", "Mock",     "N","Y","N", "Mock only",  "Read",       ""),
    ("docs_checklist.controller_signed",  "Controller Signed",          "bool", "System", "Mock",     "N","Y","N", "Mock only",  "Read",       ""),
    ("docs_checklist.confirmation_email", "Confirmation Email",         "bool", "System", "Mock",     "N","Y","N", "Mock only",  "Read",       ""),
]

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — Intake Form
# ══════════════════════════════════════════════════════════════════════════════

INTAKE_HEADERS = [
    "Section", "Form name= Attribute", "Display Label",
    "HTML Input Type", "Required?", "Python Type Stored",
    "Record Dict Key", "SQL Operation", "Notes",
]

INTAKE_ROWS = [
    # ── Section A ─────────────────────────────────────────────────────────────
    ("A", "request_type",                "Request Type",                    "select",        "Required",      "str",   "request_type",           "Write", "Options: ACH, Wire, EFT, Intra Bank Transfer"),
    ("A", "treasury_service_date",       "Treasury Service Date",           "date",          "Required",      "str",   "treasury_service_date",   "Write", ""),
    ("A", "instructions_previously_used","Previously Used Instructions",    "checkbox",      "Optional",      "bool",  "instructions_previously_used", "Write", "value='yes' → True; unchecked → False"),
    ("A", "last_used_date",              "Last Date Used",                  "date",          "Conditional",   "str",   "last_used_date",          "Write", "Required only when instructions_previously_used is checked"),
    ("A", "urgent",                      "Urgent?",                         "radio (yes/no)","Required",      "bool",  "urgent",                  "Write", "value='yes' → True"),
    ("A", "urgency_reason",              "Urgency Reason",                  "textarea",      "Conditional",   "str",   "urgency_reason",          "Write", "Shown/required when urgent=yes"),
    ("A", "prepared_by",                 "Prepared By",                     "text",          "Required",      "str",   "prepared_by",             "Write", "[AUTH] Replace with Azure AD identity"),
    ("A", "prepared_date",               "Prepared Date",                   "date",          "Auto",          "str",   "prepared_date",           "Write", "Auto-set to today via JavaScript"),
    ("A", "property_dept",               "Property / Department",           "text",          "Required",      "str",   "property_dept",           "Write", ""),
    ("A", "property_code",               "Entity / Property Code",          "text",          "Optional",      "str",   "property_code",           "Write", ""),
    ("A", "approver",                    "Approver",                        "text",          "Optional",      "str",   "approver",                "Write", "[AUTH] Will become AD user lookup"),
    ("A", "controller",                  "Controller",                      "text",          "Optional",      "str",   "controller",              "Write", "[AUTH] Will become AD user lookup"),
    # ── Section B ─────────────────────────────────────────────────────────────
    ("B", "verbal_confirmed",            "Instructions Verbally Confirmed", "checkbox",      "Optional",      "bool",  "verbal_confirmed",        "Write", "value='on'"),
    ("B", "verbal_confirmed_with_known", "Confirmed With: Known Contact",   "checkbox",      "Optional",      "bool→str","verbal_confirmed_with","Write", "Combined with _requester into one composite string on submit"),
    ("B", "verbal_confirmed_with_requester","Confirmed With: Requesting Person","checkbox",  "Optional",      "bool→str","verbal_confirmed_with","Write", "Combined with _known into one composite string on submit"),
    ("B", "verbal_contact_name",         "Verbal Contact Name",             "text",          "Conditional",   "str",   "verbal_contact_name",     "Write", ""),
    ("B", "verbal_confirm_datetime",     "Verbal Confirmation Date/Time",   "datetime-local","Conditional",   "str",   "verbal_confirm_datetime", "Write", ""),
    ("B", "avs_score",                   "AVS Score (0–100)",               "number",        "Optional",      "str",   "avs_score",               "Write", "Stored as str; warning shown if int value <90. [BANKING]"),
    ("B", "file_validation_evidence",    "Validation Evidence File",        "file",          "Optional",      "str",   "attachments.validation_evidence", "Write", "Filename recorded only. [UPLOAD]"),
    ("B", "external_source",             "Instructions from External Source","checkbox",     "Optional",      "bool",  "external_source",         "Write", ""),
    ("B", "internal_doc_not_used",       "Internal Document NOT Used",      "checkbox",      "Optional",      "bool",  "internal_doc_not_used",   "Write", ""),
    # ── Section C ─────────────────────────────────────────────────────────────
    ("C", "orig_bank_name",              "Originating Bank Name",           "text",          "Required",      "str",   "orig_bank_name",          "Write", ""),
    ("C", "orig_account_name",           "Originating Account Name",        "text",          "Required",      "str",   "orig_account_name",       "Write", ""),
    ("C", "orig_account_number",         "Originating Account Number",      "text",          "Required",      "str",   "orig_account_number",     "Write", "[RBAC] Masked on detail page"),
    ("C", "orig_routing_number",         "Originating Routing Number",      "text",          "Required",      "str",   "orig_routing_number",     "Write", "[RBAC] Masked on detail page"),
    ("C", "orig_bank_contact",           "Originating Bank Contact",        "text",          "Optional",      "str",   "orig_bank_contact",       "Write", ""),
    ("C", "notes_orig",                  "Notes — Originating Bank",        "textarea",      "Optional",      "str",   "notes_orig",              "Write", ""),
    # ── Section D ─────────────────────────────────────────────────────────────
    ("D", "recv_payee_name",             "Recipient / Payee Name",          "text",          "Required",      "str",   "recv_payee_name",         "Write", ""),
    ("D", "recv_bank_name",              "Receiving Bank Name",             "text",          "Required",      "str",   "recv_bank_name",          "Write", ""),
    ("D", "recv_account_name",           "Receiving Account Name",          "text",          "Required",      "str",   "recv_account_name",       "Write", ""),
    ("D", "recv_account_number",         "Receiving Account Number",        "text",          "Required",      "str",   "recv_account_number",     "Write", "[RBAC] Masked on detail page"),
    ("D", "recv_routing_number",         "Receiving Routing Number",        "text",          "Required",      "str",   "recv_routing_number",     "Write", "[RBAC] Masked on detail page"),
    ("D", "recv_bank_address",           "Bank / Beneficiary Address",      "textarea",      "Wire Required", "str",   "recv_bank_address",       "Write", "JS enforces required when request_type=Wire"),
    ("D", "recv_contact_name",           "Receiving Contact Name",          "text",          "Optional",      "str",   "recv_contact_name",       "Write", ""),
    ("D", "recv_contact_email",          "Receiving Contact Email",         "email",         "Optional",      "str",   "recv_contact_email",      "Write", ""),
    ("D", "recv_contact_phone",          "Receiving Contact Phone",         "tel",           "Optional",      "str",   "recv_contact_phone",      "Write", ""),
    ("D", "notes_recv",                  "Notes — Receiving Bank",          "textarea",      "Optional",      "str",   "notes_recv",              "Write", ""),
    ("D", "file_wire_ach_instructions",  "External ACH/Wire Instructions",  "file",          "Required",      "str",   "attachments.wire_ach_instructions", "Write", "Filename only. [UPLOAD]"),
    # ── Section E ─────────────────────────────────────────────────────────────
    ("E", "amount",                      "Amount",                          "number",        "Required",      "float", "amount",                  "Write", "Commas stripped before float() conversion"),
    ("E", "currency",                    "Currency",                        "select",        "Required",      "str",   "currency",                "Write", "Default: USD"),
    ("E", "payment_purpose",             "Payment Purpose / Description",   "textarea",      "Required",      "str",   "payment_purpose",         "Write", ""),
    ("E", "file_payment_support",        "Payment Support File",            "file",          "Required",      "str",   "attachments.payment_support", "Write", "Filename only. [UPLOAD]"),
    # ── Computed on submit (not form inputs) ──────────────────────────────────
    ("System", "— computed —",           "Request ID",                      "—",             "Auto",          "str",   "request_id",              "Write", "TXN-YYYY-###; random.randint(100,999)"),
    ("System", "— computed —",           "Submitted Date",                  "—",             "Auto",          "str",   "submitted_date",          "Write", "datetime.now() formatted YYYY-MM-DD"),
    ("System", "— computed —",           "Approval Tier",                   "—",             "Auto",          "str",   "approval_tier",           "Write", "get_approval_tier(amount)"),
    ("System", "— computed —",           "Status",                          "—",             "Auto",          "str",   "status",                  "Write", "Hardcoded 'Submitted' on intake"),
    ("System", "— computed —",           "Over $1M Flag",                   "—",             "Auto",          "bool",  "over_1m",                 "Write", "amount > 1,000,000"),
    ("System", "— computed —",           "Assigned Approver",               "—",             "Auto",          "str",   "assigned_approver",       "Write", "'Pending Assignment'. [WORKFLOW]"),
    ("System", "— computed —",           "Days Pending",                    "—",             "Auto",          "int",   "days_pending",            "Write", "0 on submit. [WORKFLOW]"),
    ("System", "— computed —",           "Verbal Confirmed With",           "—",             "Auto",          "str",   "verbal_confirmed_with",   "Write", "Joined string from two checkboxes"),
    ("System", "— computed —",           "Timeline",                        "—",             "Auto",          "list",  "timeline",                "Write", "First entry: 'Submitted by <prepared_by>'"),
    ("System", "— computed —",           "Comments",                        "—",             "Auto",          "list",  "comments",                "Write", "Empty list on submit"),
]

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — Dashboard
# ══════════════════════════════════════════════════════════════════════════════

DASH_HEADERS = [
    "Field Key", "Display Label", "Used In", "Data Type", "SQL Operation", "Notes",
]

DASH_ROWS = [
    # Table columns
    ("request_id",       "Request ID",             "Table — ID column + row hyperlink",            "str",   "Read", "Links to /dashboard/request/<id>"),
    ("submitted_date",   "Date",                   "Table — Date column",                          "str",   "Read", ""),
    ("request_type",     "Type",                   "Table — Type column",                          "str",   "Read", ""),
    ("property_dept",    "Property / Dept",        "Table — Property column",                      "str",   "Read", "Also searchable via filter"),
    ("prepared_by",      "Prepared By",            "Table — Prepared By column (lg+ only)",        "str",   "Read", "Hidden on small screens"),
    ("amount",           "Amount",                 "Table — Amount column",                        "float", "Read", "|currency filter; also used for >$1M row class and stat card"),
    ("status",           "Status",                 "Table — Status column (badge)",                "str",   "Read", "|status_badge_class filter drives badge color"),
    ("urgent",           "Urgent flag",            "Table — Flags column + row highlight",         "bool",  "Read", "bg-danger badge; row gets .row-urgent CSS class"),
    ("over_1m",          "Over $1M flag",          "Table — Flags column + row highlight",         "bool",  "Read", "badge-orange badge; row gets .row-over-1m CSS class"),
    ("assigned_approver","Assigned Approver",      "Table — Approver column (xl+) + filter",       "str",   "Read", ""),
    ("days_pending",     "Days Pending",           "Table — Days Pending column (xl+)",            "int",   "Read", ""),
    # Stat cards (computed counts, not direct field reads)
    ("status",           "Pending Controller Approval count","Stat card",                          "str",   "Read", "Count of records where status == 'Pending Controller Approval'"),
    ("status",           "Pending Treasury Review count",    "Stat card",                          "str",   "Read", ""),
    ("status",           "Pending Release count",            "Stat card",                          "str",   "Read", ""),
    ("status",           "Completed / Released count",       "Stat card",                          "str",   "Read", ""),
    ("urgent",           "Urgent count",                     "Stat card",                          "bool",  "Read", ""),
    ("amount",           "Over $1M count",                   "Stat card",                          "float", "Read", ""),
    # URL filter parameters (not record fields — query string params)
    ("— URL param —",    "Status",                 "Filter — Status dropdown",                     "str",   "N/A",  "?status=  Passes through to dashboard()"),
    ("— URL param —",    "Request Type",           "Filter — Type dropdown",                       "str",   "N/A",  "?request_type="),
    ("— URL param —",    "Property / Dept",        "Filter — text search (partial match)",         "str",   "N/A",  "?property=  case-insensitive substring match"),
    ("— URL param —",    "Assigned Approver",      "Filter — Approver dropdown",                   "str",   "N/A",  "?approver="),
    ("— URL param —",    "Amount Min",             "Filter — Min $ number input",                  "float", "N/A",  "?amount_min="),
    ("— URL param —",    "Amount Max",             "Filter — Max $ number input",                  "float", "N/A",  "?amount_max="),
    ("— URL param —",    "Urgent Only",            "Filter — Urgent checkbox",                     "bool",  "N/A",  "?urgent_only=1"),
    ("— URL param —",    "Over $1M Only",          "Filter — >$1M checkbox",                       "bool",  "N/A",  "?over_1m_only=1"),
]

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 4 — Request Detail
# ══════════════════════════════════════════════════════════════════════════════

DETAIL_HEADERS = [
    "Section", "Field Key", "Display Label",
    "Data Type", "Read/Write", "Masked?", "Notes",
]

DETAIL_ROWS = [
    # Section A
    ("A", "request_type",             "Request Type",                   "str",   "Read",     "No",  ""),
    ("A", "property_dept",            "Property / Department",          "str",   "Read",     "No",  ""),
    ("A", "property_code",            "Entity / Property Code",         "str",   "Read",     "No",  "Shown below property_dept"),
    ("A", "prepared_by",              "Prepared By",                    "str",   "Read",     "No",  ""),
    ("A", "prepared_date",            "Prepared Date",                  "str",   "Read",     "No",  ""),
    ("A", "approver",                 "Approver",                       "str",   "Read",     "No",  "Falls back to controller_approver key for mock data compat"),
    ("A", "controller",               "Controller",                     "str",   "Read",     "No",  ""),
    ("A", "treasury_service_date",    "Treasury Service Date",          "str",   "Read",     "No",  ""),
    ("A", "instructions_previously_used","Previously Used Instructions","bool",  "Read",     "No",  "Drives whether last_used_date or 'new instructions' notice is shown"),
    ("A", "last_used_date",           "Last Instructions Used",         "str",   "Read",     "No",  "Warning shown if blank (new/unverified instructions)"),
    ("A", "urgent",                   "Urgent?",                        "bool",  "Read",     "No",  "Also drives page header badge and urgency alert banner"),
    ("A", "urgency_reason",           "Urgency Reason",                 "str",   "Read",     "No",  "Shown in alert banner when urgent=True"),
    # Section B
    ("B", "verbal_confirmed",         "Verbally Confirmed",             "bool",  "Read",     "No",  "Checklist item — green check or red X"),
    ("B", "external_source",          "Instructions from External Source","bool","Read",     "No",  "Checklist item"),
    ("B", "internal_doc_not_used",    "Internal Document NOT Used",     "bool",  "Read",     "No",  "Checklist item"),
    ("B", "verbal_confirmed_with",    "Confirmed With",                 "str",   "Read",     "No",  "Inside verbal confirmation detail card"),
    ("B", "verbal_contact_name",      "Verbal Contact Name",            "str",   "Read",     "No",  ""),
    ("B", "verbal_confirm_datetime",  "Date / Time",                    "str",   "Read",     "No",  ""),
    ("B", "avs_score",                "AVS Score",                      "str",   "Read",     "No",  "Displayed as n/100; yellow badge if <90"),
    ("B", "docs_checklist.wf_avs_screenshot","AVS Validation Evidence", "bool",  "Read",     "No",  "Green 'Attached' or yellow 'Not attached'"),
    ("B", "attachments.validation_evidence","Validation Evidence File", "str",   "Read",     "No",  "Filename or yellow 'Not attached'"),
    # Section C
    ("C", "orig_bank_name",           "Originating Bank Name",          "str",   "Read",     "No",  ""),
    ("C", "orig_account_name",        "Originating Account Name",       "str",   "Read",     "No",  ""),
    ("C", "orig_account_number",      "Originating Account Number",     "str",   "Read",     "Yes", "Shown masked (••••••••XXXX); reveal button toggles .masked/.unmasked-value"),
    ("C", "orig_routing_number",      "Originating Routing Number",     "str",   "Read",     "Yes", "Shown masked (•••••XXXX); reveal toggle"),
    ("C", "orig_bank_contact",        "Originating Bank Contact",       "str",   "Read",     "No",  ""),
    ("C", "notes_orig",               "Notes — Originating Bank",       "str",   "Read",     "No",  ""),
    # Section D
    ("D", "recv_payee_name",          "Recipient / Payee Name",         "str",   "Read",     "No",  ""),
    ("D", "recv_bank_name",           "Receiving Bank Name",            "str",   "Read",     "No",  ""),
    ("D", "recv_account_name",        "Receiving Account Name",         "str",   "Read",     "No",  ""),
    ("D", "recv_account_number",      "Receiving Account Number",       "str",   "Read",     "Yes", "Masked; reveal toggle"),
    ("D", "recv_routing_number",      "Receiving Routing Number",       "str",   "Read",     "Yes", "Masked; reveal toggle"),
    ("D", "recv_bank_address",        "Bank / Beneficiary Address",     "str",   "Read",     "No",  ""),
    ("D", "recv_contact_name",        "Contact Name",                   "str",   "Read",     "No",  ""),
    ("D", "recv_contact_email",       "Contact Email",                  "str",   "Read",     "No",  ""),
    ("D", "recv_contact_phone",       "Contact Phone",                  "str",   "Read",     "No",  ""),
    ("D", "notes_recv",               "Notes — Receiving Bank",         "str",   "Read",     "No",  ""),
    ("D", "attachments.wire_ach_instructions","External ACH/Wire Instructions File","str","Read","No","Red 'Missing' badge if absent"),
    # Section E
    ("E", "amount",                   "Amount",                         "float", "Read",     "No",  "|currency filter; large bold display"),
    ("E", "currency",                 "Currency",                       "str",   "Read",     "No",  ""),
    ("E", "approval_tier",            "Required Approval Tier",         "str",   "Read",     "No",  "Colored tier badge (SAM/Controller/VP/CFO)"),
    ("E", "payment_purpose",          "Payment Purpose / Description",  "str",   "Read",     "No",  ""),
    ("E", "attachments.payment_support","Payment Support File",         "str",   "Read",     "No",  "Red 'Missing' badge if absent"),
    # Sidebar — Approval Actions (Write)
    ("Sidebar", "status",             "Status",                         "str",   "Write",    "No",  "POST /action — updated by approve/reject/more_info/treasury_reviewed/mark_released/mark_completed"),
    ("Sidebar", "comments",           "Comments / Notes",               "list",  "Write",    "No",  "Appended via comment textarea in action form"),
    ("Sidebar", "timeline",           "Approval Timeline",              "list",  "Write",    "No",  "New event appended on each action"),
    # Sidebar — Additional Attachments (Write)
    ("Sidebar", "extra_attachments",  "Additional Attachments",         "list",  "Write",    "No",  "POST /attach — {filename, description, uploaded_by, date}; supports multiple files per upload"),
    # Computed display-only (masked versions — not stored, generated per request)
    ("System",  "orig_account_number_masked","Masked Account Number",   "str",   "Computed", "Yes", "Generated in request_detail() via mask_account(); shows ••••••••XXXX"),
    ("System",  "orig_routing_number_masked","Masked Routing Number",   "str",   "Computed", "Yes", "Generated via mask_routing(); shows •••••XXXX"),
    ("System",  "recv_account_number_masked","Masked Recv Account",     "str",   "Computed", "Yes", ""),
    ("System",  "recv_routing_number_masked","Masked Recv Routing",     "str",   "Computed", "Yes", ""),
]

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 5 — Confirmation Page
# ══════════════════════════════════════════════════════════════════════════════

CONF_HEADERS = [
    "Field Key", "Display Label", "Data Type", "SQL Operation", "Notes",
]

CONF_ROWS = [
    ("request_id",            "Request ID",                     "str",   "Read", "Card header; also used in link to /dashboard/request/<id>"),
    ("submitted_date",        "Submitted Date",                 "str",   "Read", "Shown below request_id"),
    ("request_type",          "Request Type",                   "str",   "Read", ""),
    ("amount",                "Amount",                         "float", "Read", "|currency filter; includes currency code"),
    ("currency",              "Currency",                       "str",   "Read", "Shown alongside amount"),
    ("approval_tier",         "Required Approval Tier",         "str",   "Read", "Colored tier badge"),
    ("property_dept",         "Property / Department",          "str",   "Read", ""),
    ("prepared_by",           "Prepared By",                    "str",   "Read", ""),
    ("treasury_service_date", "Treasury Service Date",          "str",   "Read", ""),
    ("urgent",                "Urgent?",                        "bool",  "Read", "Shows alert-danger banner when True"),
    ("urgency_reason",        "Urgency Reason",                 "str",   "Read", "Inside urgent alert banner"),
    ("over_1m",               "Over $1M",                       "bool",  "Read", "Shows high-value-banner when True"),
    ("recv_payee_name",       "Recipient / Payee Name",         "str",   "Read", ""),
]

# ══════════════════════════════════════════════════════════════════════════════
#  Build workbook
# ══════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    write_sheet(
        wb, "Master Field Reference",
        MASTER_HEADERS, MASTER_ROWS,
        col_widths=[28, 30, 10, 8, 9, 9, 7, 12, 12, 14, 40],
        section_col=3,
        rw_col=9,
    )

    write_sheet(
        wb, "Intake Form",
        INTAKE_HEADERS, INTAKE_ROWS,
        col_widths=[8, 28, 30, 16, 14, 18, 30, 14, 40],
        section_col=0,
        rw_col=7,
    )

    write_sheet(
        wb, "Dashboard",
        DASH_HEADERS, DASH_ROWS,
        col_widths=[22, 28, 42, 10, 14, 40],
        rw_col=4,
    )

    write_sheet(
        wb, "Request Detail",
        DETAIL_HEADERS, DETAIL_ROWS,
        col_widths=[8, 30, 30, 10, 10, 8, 45],
        section_col=0,
        rw_col=4,
    )

    write_sheet(
        wb, "Confirmation",
        CONF_HEADERS, CONF_ROWS,
        col_widths=[26, 30, 12, 14, 48],
        rw_col=3,
    )

    out = "E-Transaction Field Map.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
